from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import io
import os
#from functools import wraps
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime

# --- CONFIGURA TU USUARIO Y CONTRASEÑA AQUÍ ---

USERNAME = "COSTO"
PASSWORD = "2525"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "..", "fronthead")

app = Flask(__name__, template_folder="../fronthead")
app.secret_key = "CLAVE_SECRETA_SEGURA"

#pdf
def generar_pdf(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items):

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)

    # FECHA DE HOY
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")

    # -----------------------------
    # LOGO (arriba a la derecha)
    # -----------------------------
    try:
        logo = ImageReader("logo.png")

        # Posición: margen derecho restando el ancho del logo
        logo_width = 139.2
        logo_height = 35
        x_logo = 560 - logo_width   # 560 es el ancho útil de la hoja (letter con margen)
        y_logo = 740                # altura para que quede arriba

        pdf.drawImage(
            logo,
            x_logo,
            y_logo,
            width=logo_width,
            height=logo_height,
            mask='auto'
        )
    except:
        pass

    # -----------------------------
    # TÍTULO (debajo del logo)
    # -----------------------------
    pdf.setFont("Helvetica-Bold", 16)

    # Ajuste automático para centrar el título
    titulo = "Orden de servicio"
    tamano_titulo = pdf.stringWidth(titulo, "Helvetica-Bold", 16)
    x_titulo = (612 - tamano_titulo) / 2   # 612 = ancho total hoja carta

    pdf.drawString(x_titulo, 720, titulo)

    # Fecha bajo el título, alineada a la derecha
    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, 740, f"Fecha: {fecha_hoy}")

    # Línea gris
    pdf.setStrokeColorRGB(0.6, 0.6, 0.6)
    pdf.setLineWidth(1)
    pdf.line(50, 700, 560, 700)

    # -----------------------------
    # DATOS DEL CLIENTE
    # -----------------------------
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(50, 680, f"Orden N°: {num_orden}")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, 655, f"Nombre: {nombre}")
    pdf.drawString(300, 655, f"Cédula: {cedula}")

    pdf.drawString(50, 630, f"Empresa: {empresa}")
    pdf.drawString(300, 630, f"Vehículo: {vehiculo}")

    pdf.drawString(50, 605, f"Teléfono: {telefono}")

    # Línea divisora
    pdf.line(50, 590, 560, 590)

    # -----------------------------
    # LISTADO DE ITEMS
    # -----------------------------
    y = 570

    pdf.setFont("Helvetica-Bold", 13)
    pdf.drawString(50, y, "Servicios de llantas Solicitados:")
    y -= 25

    pdf.setFont("Helvetica", 11)

    for i, item in enumerate(items, start=1):

        if y < 60:
            pdf.showPage()
            y = 750
            pdf.setFont("Helvetica", 11)

        texto = (
            f"{i}. Marca: {item['Marca']}  |  "
            f"Ref: {item['Referencia']}  |  "
            f"Serie: {item['Serie']}  |  "
            f"Tipo: {item['Servicio']}"
        )

        pdf.drawString(50, y, texto)
        y -= 20

    pdf.save()
    buffer.seek(0)
    return buffer

# configurar googlesheets

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
import json
CREDS_FILE = json.loads(os.getenv("CREDS_JSON"))
SPREADSHEET_ID = "1iL1FLRb42b-mBZrSc2VVoC7NQJoQ8f420yrOaugrpKo"


def guardar_google_sheets(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items):

    # autenticacion google sheets
    creds = Credentials.from_service_account_info(CREDS_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)

    # abrir hoja
    sheet = client.open_by_key(SPREADSHEET_ID).sheet1

    # guardar items
    for index, item in enumerate(items, start=1):
        fila = [
            fecha_hoy,
            num_orden,
            index,
            nombre,
            cedula,
            empresa,
            vehiculo,
            telefono,
            item["Marca"],
            item["Referencia"],
            item["Serie"],
            item["Servicio"],
            item["Observacion"],
        ]
        sheet.append_row(fila)

def obtener_numero_orden():
    """
    Obtiene el siguiente número de orden leyendo 
    el último valor registrado en Google Sheets.
    """
    creds = Credentials.from_service_account_info(CREDS_FILE, scopes=SCOPES)
    client = gspread.authorize(creds)

    sheet = client.open_by_key(SPREADSHEET_ID).sheet1
    columna_orden = sheet.col_values(2)  # columna B completa

    # Si no hay registros, empezamos en 1
    if len(columna_orden) <= 1:
        return 1

    try:
        ultimo = int(columna_orden[-1])
    except:
        ultimo = 0

    return ultimo + 1


def guardar_excel(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items):

    excel_file = "registros.xlsx"

    if not os.path.exists(excel_file):
        # Crear un nuevo archivo Excel y agregar encabezados
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"
        ws.append(
            [
                "fecha_hoy",
                "Num_Orden",
                "Item",
                "Nombre",
                "Cedula",
                "Empresa",
                "Vehiculo",
                "Telefono",
                "Marca",
                "Referencia",
                "Serie",
                "Servicio",
                "Observacion",
            ]
        )
    else:
        # Abrir archivo existente y agregar el registro
        wb = load_workbook(excel_file)
        ws = wb.active

    # guardar cada item como fila
    for index, item in enumerate(items, start=1):
        ws.append([
                fecha_hoy,
                num_orden,
                index,
                nombre,
                cedula,
                empresa,
                vehiculo,
                telefono,
                item["Marca"],
                item["Referencia"],
                item["Serie"],
                item["Servicio"],
                item["Observacion"],
            ])

    wb.save(excel_file)

#login
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = request.form["username"]
        pwd = request.form["password"]

        if user == USERNAME and pwd == PASSWORD:
            session["logged"] = True
            return redirect(url_for("index"))
        else:
            return "Usuario o contraseña incorrectos"

    return render_template("login.html")



# ruta principal - formualrio

@app.route("/")
# @require_auth
def index():
    if not session.get("logged"):
        return redirect(url_for("login"))
    num_orden = obtener_numero_orden()
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    return render_template("index.html", num_orden=num_orden, fecha_hoy=fecha_hoy)

@app.route("/ver_pdf")
def ver_pdf():
    file = request.args.get("file")
    if not file or not os.path.exists(file):
        return "Archivo no encontrado", 404

    return send_file(file, mimetype="application/pdf")

# ruta formulario
@app.route("/register", methods=["POST"])
# @require_auth
def register():
    
    if not session.get("logged"):
        return redirect(url_for("login"))
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    num_orden = obtener_numero_orden()
    nombre = request.form["Nombre"]
    cedula = request.form["Cedula"]
    empresa = request.form["Empresa"]
    vehiculo = request.form["Vehiculo"]
    telefono = request.form["Telefono"]

    # datos tabla

    marca = request.form.getlist("Marca[]")
    referencia = request.form.getlist("Referencias[]")
    serie = request.form.getlist("Serie[]")
    servicio = request.form.getlist("Tipos[]")
    observacion = request.form.getlist("Observacion[]")

    # combinar items

    items = []
    for s, c, v, st, o in zip(marca, referencia, serie, servicio, observacion):
        items.append({"Marca": s, "Referencia": c, "Serie": v, "Servicio": st, "Observacion": o})

    # guardar excel
    guardar_excel(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items)
    # Guardar en Google Sheets
    guardar_google_sheets(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items)
    # Generar PDF
    pdf_buffer = generar_pdf(fecha_hoy, num_orden, nombre, cedula, empresa, vehiculo, telefono, items)

    # Guardar el PDF temporalmente con nombre único
    pdf_path = f"Recibo_{num_orden}.pdf"
    with open(pdf_path, "wb") as f:
        f.write(bytes(pdf_buffer.getbuffer()))

    # Enviar página de descarga + redirección
    return render_template("vista_pdf.html",
                           pdf_path=pdf_path,
                           num_orden=num_orden)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)









